"""Product catalog — CRUD, bulk import, AI extraction, deal matching."""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID

from fastapi import HTTPException, UploadFile
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.client import Client
from app.models.crm_lead import CrmLead
from app.models.product import Product, ProductImportJob
from app.schemas.product import ProductCreate, ProductUpdate
from app.services.ai_service import _extract_json, _validate_api_key, get_openai

logger = logging.getLogger(__name__)

IMPORT_STATUSES = frozenset({"pending", "processing", "completed", "failed"})
IMPORT_SOURCES = frozenset({"csv", "xlsx", "pdf", "text"})

_HEADER_ALIASES: dict[str, str] = {
    "name": "name",
    "product": "name",
    "product name": "name",
    "product_name": "name",
    "title": "name",
    "sku": "sku",
    "code": "sku",
    "product code": "sku",
    "category": "category",
    "cat": "category",
    "type": "category",
    "description": "description",
    "desc": "description",
    "details": "description",
    "moq": "moq",
    "min order": "moq",
    "minimum order": "moq",
    "unit_price": "unit_price",
    "unit price": "unit_price",
    "price": "unit_price",
    "currency": "currency",
}

_EXTRACT_SYSTEM = """\
You extract structured product catalog rows from supplier catalog text.
Return ONLY JSON:
{
  "products": [
    {
      "name": "string (required)",
      "sku": "string or null",
      "category": "string or null",
      "description": "string or null",
      "moq": integer or null,
      "unit_price": number or null,
      "currency": "USD|UZS|CNY|...",
      "attributes": { "key": "value" }
    }
  ]
}
Rules:
- Extract every distinct product you can identify
- Normalize MOQ and price to numbers when possible
- Keep specs in attributes object
- Max 50 products per response
"""

_MATCH_SYSTEM = """\
You match CRM lead interest to a client's product catalog for B2B sales in Uzbekistan.
Return ONLY JSON:
{
  "matches": [
    {
      "product_id": "uuid string from catalog",
      "confidence": 0.0 to 1.0,
      "reason": "short explanation"
    }
  ]
}
Rules:
- Return up to 5 best matches sorted by confidence
- Only use product_ids from the provided catalog
- confidence reflects fit to lead interest, notes, and company context
- Never invent products
"""


def _normalize_header(key: str) -> str:
    return _HEADER_ALIASES.get(key.strip().lower().replace("-", " "), key.strip().lower())


def _parse_decimal(val: Any) -> Decimal | None:
    if val is None or val == "":
        return None
    try:
        cleaned = re.sub(r"[^\d.,-]", "", str(val).replace(",", ""))
        return Decimal(cleaned) if cleaned else None
    except (InvalidOperation, ValueError):
        return None


def _parse_int(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(float(str(val).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _map_row(raw: dict[str, Any]) -> dict[str, Any] | None:
    mapped: dict[str, Any] = {}
    attrs: dict[str, Any] = {}
    for key, val in raw.items():
        if val is None or str(val).strip() == "":
            continue
        norm = _normalize_header(str(key))
        if norm in ("name", "sku", "category", "description", "moq", "unit_price", "currency"):
            mapped[norm] = val
        else:
            attrs[str(key)] = val
    name = mapped.get("name")
    if not name or not str(name).strip():
        return None
    mapped["name"] = str(name).strip()
    mapped["moq"] = _parse_int(mapped.get("moq"))
    mapped["unit_price"] = _parse_decimal(mapped.get("unit_price"))
    if attrs:
        mapped["attributes_json"] = attrs
    return mapped


class ProductCatalogService:
    @staticmethod
    def _serialize(product: Product, *, company_name: str | None = None) -> dict[str, Any]:
        return {
            "id": product.id,
            "client_id": product.client_id,
            "name": product.name,
            "sku": product.sku,
            "category": product.category,
            "description": product.description,
            "moq": product.moq,
            "unit_price": product.unit_price,
            "currency": product.currency,
            "attributes_json": product.attributes_json,
            "images_json": product.images_json,
            "active": product.active,
            "created_at": product.created_at,
            "company_name": company_name,
        }

    @staticmethod
    async def _ensure_client(db: AsyncSession, client_id: UUID) -> Client:
        result = await db.execute(select(Client).where(Client.id == client_id))
        client = result.scalar_one_or_none()
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        return client

    @staticmethod
    async def list_products(
        db: AsyncSession,
        *,
        client_id: UUID | None = None,
        category: str | None = None,
        search: str | None = None,
        active: bool | None = True,
        skip: int = 0,
        limit: int = 50,
    ) -> dict[str, Any]:
        query = (
            select(Product, Client.company_name)
            .join(Client, Client.id == Product.client_id)
            .order_by(Product.created_at.desc())
        )
        count_q = select(func.count()).select_from(Product)
        if client_id:
            query = query.where(Product.client_id == client_id)
            count_q = count_q.where(Product.client_id == client_id)
        if category:
            query = query.where(Product.category == category)
            count_q = count_q.where(Product.category == category)
        if active is not None:
            query = query.where(Product.active == active)
            count_q = count_q.where(Product.active == active)
        if search:
            term = f"%{search.strip()}%"
            filt = or_(
                Product.name.ilike(term),
                Product.sku.ilike(term),
                Product.description.ilike(term),
                Product.category.ilike(term),
            )
            query = query.where(filt)
            count_q = count_q.where(filt)

        total = int(await db.scalar(count_q) or 0)
        result = await db.execute(query.offset(skip).limit(limit))
        items = [
            ProductCatalogService._serialize(p, company_name=cn)
            for p, cn in result.all()
        ]
        logger.info("[Product Catalog] listed: total=%s returned=%s", total, len(items))
        return {"items": items, "total": total}

    @staticmethod
    async def get_product(db: AsyncSession, product_id: UUID) -> dict[str, Any]:
        result = await db.execute(
            select(Product, Client.company_name)
            .join(Client, Client.id == Product.client_id)
            .where(Product.id == product_id)
        )
        row = result.one_or_none()
        if not row:
            raise HTTPException(status_code=404, detail="Product not found")
        product, company_name = row
        return ProductCatalogService._serialize(product, company_name=company_name)

    @staticmethod
    async def create_product(db: AsyncSession, data: ProductCreate) -> dict[str, Any]:
        await ProductCatalogService._ensure_client(db, data.client_id)
        product = Product(
            client_id=data.client_id,
            name=data.name.strip(),
            sku=data.sku.strip() if data.sku else None,
            category=data.category.strip() if data.category else None,
            description=data.description,
            moq=data.moq,
            unit_price=data.unit_price,
            currency=(data.currency or "USD").upper(),
            attributes_json=data.attributes_json,
            images_json=data.images_json,
            active=data.active,
        )
        db.add(product)
        await db.commit()
        await db.refresh(product)
        logger.info("[Product Catalog] created: product=%s client=%s", product.id, data.client_id)
        return await ProductCatalogService.get_product(db, product.id)

    @staticmethod
    async def update_product(
        db: AsyncSession,
        product_id: UUID,
        data: ProductUpdate,
    ) -> dict[str, Any]:
        result = await db.execute(select(Product).where(Product.id == product_id))
        product = result.scalar_one_or_none()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        for field, value in data.model_dump(exclude_unset=True).items():
            if field == "currency" and value:
                value = value.upper()
            setattr(product, field, value)
        await db.commit()
        logger.info("[Product Catalog] updated: product=%s", product_id)
        return await ProductCatalogService.get_product(db, product_id)

    @staticmethod
    async def delete_product(db: AsyncSession, product_id: UUID) -> None:
        result = await db.execute(select(Product).where(Product.id == product_id))
        product = result.scalar_one_or_none()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        await db.delete(product)
        await db.commit()
        logger.info("[Product Catalog] deleted: product=%s", product_id)

    @staticmethod
    def _parse_csv(content: bytes) -> list[dict[str, Any]]:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows: list[dict[str, Any]] = []
        for row in reader:
            mapped = _map_row(row)
            if mapped:
                rows.append(mapped)
        return rows

    @staticmethod
    def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="Excel import requires openpyxl on the server",
            ) from exc
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h or "").strip() for h in next(rows_iter)]
        except StopIteration:
            return []
        rows: list[dict[str, Any]] = []
        for values in rows_iter:
            raw = {headers[i]: values[i] for i in range(len(headers)) if headers[i]}
            mapped = _map_row(raw)
            if mapped:
                rows.append(mapped)
        return rows

    @staticmethod
    def _parse_pdf(content: bytes) -> str:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="PDF import requires pypdf on the server",
            ) from exc
        reader = PdfReader(io.BytesIO(content))
        parts = []
        for page in reader.pages[:30]:
            parts.append(page.extract_text() or "")
        return "\n".join(parts).strip()

    @staticmethod
    async def _ai_extract_products(text: str) -> list[dict[str, Any]]:
        snippet = text[:12000]
        if not snippet.strip():
            return []
        try:
            _validate_api_key()
            openai = get_openai()
            response = await openai.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": _EXTRACT_SYSTEM},
                    {"role": "user", "content": f"Extract products from this catalog text:\n\n{snippet}"},
                ],
                temperature=0.2,
            )
            parsed = _extract_json(response.choices[0].message.content or "{}")
            products = parsed.get("products") or []
            out: list[dict[str, Any]] = []
            for p in products:
                if not isinstance(p, dict) or not p.get("name"):
                    continue
                attrs = p.get("attributes") if isinstance(p.get("attributes"), dict) else {}
                out.append({
                    "name": str(p["name"]).strip(),
                    "sku": p.get("sku"),
                    "category": p.get("category"),
                    "description": p.get("description"),
                    "moq": _parse_int(p.get("moq")),
                    "unit_price": _parse_decimal(p.get("unit_price")),
                    "currency": (p.get("currency") or "USD").upper(),
                    "attributes_json": attrs or None,
                })
            logger.info("[Product Import] AI extracted: count=%s", len(out))
            return out
        except ValueError:
            logger.info("[Product Import] AI unavailable — using heuristic extraction")
            return ProductCatalogService._heuristic_extract(snippet)

    @staticmethod
    def _heuristic_extract(text: str) -> list[dict[str, Any]]:
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        products: list[dict[str, Any]] = []
        for ln in lines[:100]:
            if len(ln) < 4 or ln.startswith("#"):
                continue
            price = None
            m = re.search(r"(\d[\d,.\s]{0,12})\s*(USD|UZS|CNY|\$)", ln, re.I)
            if m:
                price = _parse_decimal(m.group(1))
            products.append({
                "name": ln[:200],
                "sku": None,
                "category": None,
                "description": ln,
                "moq": None,
                "unit_price": price,
                "currency": (m.group(2).upper() if m else "USD").replace("$", "USD"),
                "attributes_json": None,
            })
        return products[:20]

    @staticmethod
    async def _bulk_create(
        db: AsyncSession,
        client_id: UUID,
        rows: list[dict[str, Any]],
    ) -> tuple[int, int, list[str]]:
        imported = 0
        skipped = 0
        errors: list[str] = []
        for i, row in enumerate(rows, start=1):
            try:
                name = str(row.get("name", "")).strip()
                if not name:
                    skipped += 1
                    continue
                product = Product(
                    client_id=client_id,
                    name=name,
                    sku=str(row["sku"]).strip() if row.get("sku") else None,
                    category=str(row["category"]).strip() if row.get("category") else None,
                    description=row.get("description"),
                    moq=row.get("moq"),
                    unit_price=row.get("unit_price"),
                    currency=str(row.get("currency") or "USD").upper(),
                    attributes_json=row.get("attributes_json"),
                    images_json=row.get("images_json"),
                    active=True,
                )
                db.add(product)
                imported += 1
            except Exception as exc:
                skipped += 1
                errors.append(f"Row {i}: {exc}")
        await db.commit()
        return imported, skipped, errors

    @staticmethod
    async def import_catalog(
        db: AsyncSession,
        *,
        client_id: UUID,
        source_type: str,
        file: UploadFile | None = None,
        catalog_text: str | None = None,
    ) -> dict[str, Any]:
        if source_type not in IMPORT_SOURCES:
            raise HTTPException(status_code=400, detail=f"Invalid source_type: {source_type}")
        await ProductCatalogService._ensure_client(db, client_id)

        job = ProductImportJob(
            client_id=client_id,
            source_type=source_type,
            source_file=file.filename if file else None,
            status="processing",
        )
        db.add(job)
        await db.flush()
        logger.info(
            "[Product Import] started: job=%s client=%s source=%s file=%s",
            job.id, client_id, source_type, job.source_file,
        )

        errors: list[str] = []
        rows: list[dict[str, Any]] = []
        try:
            if source_type == "text":
                if not catalog_text or not catalog_text.strip():
                    raise HTTPException(status_code=400, detail="catalog_text is required for text import")
                rows = await ProductCatalogService._ai_extract_products(catalog_text)
            else:
                if not file:
                    raise HTTPException(status_code=400, detail="file is required for file import")
                content = await file.read()
                if source_type == "csv":
                    rows = ProductCatalogService._parse_csv(content)
                elif source_type == "xlsx":
                    rows = ProductCatalogService._parse_xlsx(content)
                elif source_type == "pdf":
                    text = ProductCatalogService._parse_pdf(content)
                    rows = await ProductCatalogService._ai_extract_products(text)
                if source_type in ("csv", "xlsx") and not rows:
                    errors.append("No valid product rows found in file")

            imported, skipped, row_errors = await ProductCatalogService._bulk_create(db, client_id, rows)
            errors.extend(row_errors)
            job.status = "completed"
            job.result_json = {
                "imported": imported,
                "skipped": skipped,
                "parsed_rows": len(rows),
                "errors": errors,
            }
            await db.commit()
            await db.refresh(job)
            logger.info(
                "[Product Import] completed: job=%s imported=%s skipped=%s",
                job.id, imported, skipped,
            )
        except HTTPException:
            job.status = "failed"
            job.result_json = {"errors": errors}
            await db.commit()
            raise
        except Exception as exc:
            job.status = "failed"
            job.result_json = {"errors": errors + [str(exc)]}
            await db.commit()
            logger.warning("[Product Import] failed: job=%s error=%s", job.id, exc)
            raise HTTPException(status_code=500, detail=f"Import failed: {exc}") from exc

        return {
            "job": {
                "id": job.id,
                "client_id": job.client_id,
                "source_type": job.source_type,
                "source_file": job.source_file,
                "status": job.status,
                "result_json": job.result_json,
                "created_at": job.created_at,
            },
            "imported": job.result_json.get("imported", 0),
            "skipped": job.result_json.get("skipped", 0),
            "errors": errors,
        }

    @staticmethod
    def _lead_context(lead: CrmLead) -> str:
        parts = [
            f"Lead: {lead.name}",
            f"Company: {lead.company or ''}",
            f"Interest: {lead.interest or ''}",
            f"Notes: {lead.notes or ''}",
            f"Status: {lead.status}",
        ]
        return "\n".join(p for p in parts if p.strip())

    @staticmethod
    def _keyword_score(lead_text: str, product: Product) -> tuple[float, str]:
        text = lead_text.lower()
        tokens = [t for t in re.split(r"\W+", text) if len(t) > 2]
        blob = " ".join(
            filter(None, [product.name, product.category, product.description, product.sku])
        ).lower()
        if not blob:
            return 0.0, "No product text to match"
        hits = sum(1 for t in tokens if t in blob)
        if hits == 0:
            return 0.0, "No keyword overlap"
        score = min(0.95, 0.35 + hits * 0.12)
        return score, f"Matched {hits} keyword(s) from lead interest"

    @staticmethod
    async def match_lead(db: AsyncSession, lead_id: UUID) -> dict[str, Any]:
        result = await db.execute(
            select(CrmLead)
            .options(selectinload(CrmLead.client))
            .where(CrmLead.id == lead_id)
        )
        lead = result.scalar_one_or_none()
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")

        products_r = await db.execute(
            select(Product)
            .where(Product.client_id == lead.client_id, Product.active.is_(True))
            .order_by(Product.name)
            .limit(200)
        )
        products = list(products_r.scalars().all())
        context = ProductCatalogService._lead_context(lead)
        logger.info(
            "[Deal Matching] lead=%s client=%s products=%s",
            lead_id, lead.client_id, len(products),
        )

        if not products:
            return {
                "lead_id": lead.id,
                "lead_name": lead.name,
                "query_context": context,
                "matches": [],
                "demo_mode": False,
            }

        demo_mode = False
        matches: list[dict[str, Any]] = []

        try:
            _validate_api_key()
            catalog = [
                {
                    "product_id": str(p.id),
                    "name": p.name,
                    "sku": p.sku,
                    "category": p.category,
                    "description": (p.description or "")[:300],
                    "unit_price": float(p.unit_price) if p.unit_price is not None else None,
                    "currency": p.currency,
                    "moq": p.moq,
                }
                for p in products
            ]
            openai = get_openai()
            response = await openai.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": _MATCH_SYSTEM},
                    {
                        "role": "user",
                        "content": (
                            f"Lead context:\n{context}\n\n"
                            f"Product catalog JSON:\n{json.dumps(catalog, ensure_ascii=False)}"
                        ),
                    },
                ],
                temperature=0.2,
            )
            parsed = _extract_json(response.choices[0].message.content or "{}")
            by_id = {str(p.id): p for p in products}
            for m in (parsed.get("matches") or [])[:5]:
                pid = str(m.get("product_id", ""))
                product = by_id.get(pid)
                if not product:
                    continue
                conf = float(m.get("confidence") or 0)
                matches.append({
                    "product_id": product.id,
                    "name": product.name,
                    "sku": product.sku,
                    "category": product.category,
                    "unit_price": product.unit_price,
                    "currency": product.currency,
                    "confidence": max(0.0, min(1.0, conf)),
                    "reason": str(m.get("reason") or "AI match"),
                })
        except ValueError:
            demo_mode = True
            scored = []
            for p in products:
                score, reason = ProductCatalogService._keyword_score(context, p)
                if score > 0:
                    scored.append((score, reason, p))
            scored.sort(key=lambda x: x[0], reverse=True)
            for score, reason, p in scored[:5]:
                matches.append({
                    "product_id": p.id,
                    "name": p.name,
                    "sku": p.sku,
                    "category": p.category,
                    "unit_price": p.unit_price,
                    "currency": p.currency,
                    "confidence": round(score, 2),
                    "reason": reason,
                })

        matches.sort(key=lambda x: x["confidence"], reverse=True)
        logger.info("[Deal Matching] lead=%s matches=%s demo=%s", lead_id, len(matches), demo_mode)
        return {
            "lead_id": lead.id,
            "lead_name": lead.name,
            "query_context": context,
            "matches": matches,
            "demo_mode": demo_mode,
        }

    @staticmethod
    async def list_categories(db: AsyncSession, client_id: UUID | None = None) -> list[str]:
        query = select(Product.category).where(Product.category.isnot(None)).distinct()
        if client_id:
            query = query.where(Product.client_id == client_id)
        result = await db.execute(query.order_by(Product.category))
        return [c for c in result.scalars().all() if c]
